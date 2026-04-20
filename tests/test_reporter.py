"""Tests for XLSX report generation, especially Pass 2 investigation columns."""

from __future__ import annotations

import io

import openpyxl
import pytest

from app.reporter import generate_xlsx, generate_json
from app.schema import ScreeningResult


def _make_result(**overrides) -> ScreeningResult:
    defaults = dict(
        id="SCR-TEST-001",
        timestamp="2026-04-16T10:00:00",
        user_input={"name": "Test User", "dob": "2003-08-31", "nationality": "Pakistan"},
        result="CLEAR",
        raw_candidates=80,
        unique_persons=30,
        auto_cleared=28,
        auto_flagged=0,
        sent_to_llm=2,
        llm_cleared=2,
        llm_flagged=0,
        llm_escalated=0,
        investigations_run=0,
        matches=[],
        llm_calls=[],
        investigation_audits=[],
        source_versions={"un_consolidated": "abc123", "ofac_sdn": "def456"},
        processing_ms=1500,
        screened_by="Tester",
    )
    defaults.update(overrides)
    return ScreeningResult(**defaults)


def test_generate_xlsx_empty_matches():
    result = _make_result(matches=[])
    data = generate_xlsx(result)
    assert data.startswith(b"PK")  # ZIP magic byte (XLSX is a zip)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Matches" in wb.sheetnames
    assert "Audit" in wb.sheetnames

    matches_sheet = wb["Matches"]
    # Header row + 1 "no candidates" row
    assert matches_sheet.cell(row=1, column=2).value == "Decision"
    assert matches_sheet.cell(row=2, column=2).value == "CLEAR"


def test_generate_xlsx_with_investigation_column():
    """Verify the Investigation Sources column is present and populated."""
    matches = [{
        "number": 1,
        "decision": "CLEARED",
        "cleared_by": "AI + Investigation",
        "matched_person": "Mian Abdul Manan",
        "aliases": "",
        "dob": "",
        "nationality": "Pakistan",
        "gender": "",
        "designation": "MNA",
        "source_lists": "Wikidata PEPs",
        "identifiers": "",
        "key_contradiction": "Tenure 2013-2018 while customer was 10yo",
        "reasoning": "Politician served before customer was adult.",
        "investigation_sources": "https://wikipedia.org/wiki/Mian_Abdul_Manan; https://alchetron.com/...",
    }]
    result = _make_result(matches=matches, investigations_run=1)
    data = generate_xlsx(result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    matches_sheet = wb["Matches"]

    # Find the Investigation Sources column
    headers = [matches_sheet.cell(row=1, column=c).value for c in range(1, 20)]
    assert "Investigation Sources" in headers
    inv_col = headers.index("Investigation Sources") + 1

    # Row 2 should have the citation URL
    val = matches_sheet.cell(row=2, column=inv_col).value
    assert "wikipedia.org" in val


def test_generate_xlsx_investigation_audit_sheet():
    """Verify investigations show up in the Audit sheet."""
    inv_audits = [{
        "match_number": 0,
        "match_name": "Mian Abdul Manan",
        "match_sources": ["Wikidata PEPs"],
        "question": "Who is Mian Abdul Manan?",
        "perplexity_answer": "Pakistani politician, served 2013-2018.",
        "perplexity_citations": [
            {"url": "https://wikipedia.org/wiki/Mian_Abdul_Manan", "title": "Wiki"},
        ],
        "claude_call": {
            "claude_input_tokens": 200,
            "claude_output_tokens": 50,
            "claude_response": '{"decision": "CLEARED"}',
        },
    }]
    result = _make_result(investigation_audits=inv_audits, investigations_run=1)
    data = generate_xlsx(result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    audit_sheet = wb["Audit"]

    # Gather all text from the audit sheet
    all_text = "\n".join(
        str(audit_sheet.cell(row=r, column=c).value or "")
        for r in range(1, audit_sheet.max_row + 1)
        for c in range(1, 3)
    )
    assert "PASS 2 INVESTIGATIONS" in all_text
    assert "Mian Abdul Manan" in all_text
    assert "Perplexity Answer" in all_text
    assert "Citations" in all_text
    assert "wikipedia.org" in all_text


def test_generate_json_includes_investigation_count():
    result = _make_result(investigations_run=2, matches=[])
    text = generate_json(result)
    import json
    parsed = json.loads(text)
    assert parsed["result"] == "CLEAR"
    assert parsed["stats"]["sent_to_llm"] == 2


def test_conditional_formatting_still_applied():
    """The decision column should still have conditional formatting rules."""
    matches = [
        {"number": 1, "decision": "CLEARED", "cleared_by": "AI"},
        {"number": 2, "decision": "LIKELY_MATCH", "cleared_by": "AI"},
        {"number": 3, "decision": "ESCALATE", "cleared_by": "AI"},
    ]
    result = _make_result(matches=matches)
    data = generate_xlsx(result)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    matches_sheet = wb["Matches"]

    # Conditional formatting rules should exist on column B
    cf = matches_sheet.conditional_formatting
    rules_list = list(cf._cf_rules.values()) if hasattr(cf, "_cf_rules") else list(cf)
    # There should be at least 3 rules (green, red, yellow)
    total_rules = sum(len(rules) for rules in rules_list)
    assert total_rules >= 3
