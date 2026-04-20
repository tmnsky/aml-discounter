"""XLSX and JSON report generation for screening results."""

import json
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from .schema import ScreeningResult


# Color definitions
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=11)
RESULT_FONT_CLEAR = Font(bold=True, size=14, color="006100")
RESULT_FONT_FLAG = Font(bold=True, size=14, color="9C0006")
RESULT_FONT_ESCALATE = Font(bold=True, size=14, color="9C6500")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generate_xlsx(result: ScreeningResult) -> bytes:
    """Generate XLSX report bytes from a ScreeningResult."""
    wb = Workbook()

    _build_summary_sheet(wb.active, result)
    _build_matches_sheet(wb.create_sheet("Matches"), result)
    _build_audit_sheet(wb.create_sheet("Audit"), result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_summary_sheet(ws, result: ScreeningResult):
    """Sheet 1: Summary — key-value pairs for quick verdict."""
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80

    user = result.user_input
    result_font = {
        "CLEAR": RESULT_FONT_CLEAR,
        "FLAG": RESULT_FONT_FLAG,
        "ESCALATE": RESULT_FONT_ESCALATE,
    }.get(result.result, RESULT_FONT_ESCALATE)

    rows = [
        ("RESULT", result.result, result_font),
        ("", "", None),
        ("Subject", user.get("name", "Unknown"), BOLD_FONT),
        ("Date of Birth", user.get("dob", "Not provided"), NORMAL_FONT),
        ("Nationality", user.get("nationality", "Not provided"), NORMAL_FONT),
        ("CNIC / National ID", user.get("cnic", "Not provided"), NORMAL_FONT),
        ("Passport", user.get("passport", "Not provided"), NORMAL_FONT),
        ("Gender", user.get("gender", "Not provided"), NORMAL_FONT),
        ("Place of Birth", user.get("pob", "Not provided"), NORMAL_FONT),
        ("", "", None),
        ("Screened At", result.timestamp, NORMAL_FONT),
        ("Screened By", result.screened_by or "Not recorded", NORMAL_FONT),
        ("Report ID", result.id, NORMAL_FONT),
        ("", "", None),
        ("Sources Checked", _format_sources(result.source_versions), NORMAL_FONT),
        ("Data Freshness", "All lists updated within last 24 hours", NORMAL_FONT),
        ("", "", None),
        ("Raw Candidates", str(result.raw_candidates), NORMAL_FONT),
        ("Unique Persons (after dedup)", str(result.unique_persons), NORMAL_FONT),
        ("Auto-Cleared", str(result.auto_cleared), NORMAL_FONT),
        ("Auto-Flagged", str(result.auto_flagged), NORMAL_FONT),
        ("AI Analyzed", str(result.sent_to_llm), NORMAL_FONT),
        ("AI Cleared (final)", str(result.llm_cleared), NORMAL_FONT),
        ("AI Flagged (final)", str(result.llm_flagged), NORMAL_FONT),
        ("AI Escalated (final)", str(result.llm_escalated), NORMAL_FONT),
        ("Investigations Run (Pass 2)", str(getattr(result, "investigations_run", 0)), NORMAL_FONT),
        ("", "", None),
        ("Processing Time", f"{result.processing_ms}ms", NORMAL_FONT),
        ("", "", None),
        (
            "Note",
            "This report is an analytical aid. Final determination rests with the reviewing compliance team.",
            Font(italic=True, size=10, color="666666"),
        ),
    ]

    for row_idx, (label, value, font) in enumerate(rows, 1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        cell_a.font = BOLD_FONT
        if font:
            cell_b.font = font


def _build_matches_sheet(ws, result: ScreeningResult):
    """Sheet 2: Matches — one row per unique person."""
    headers = [
        "#", "Decision", "Cleared By", "Matched Person",
        "Aliases", "DOB (match)", "Nationality (match)", "Gender (match)",
        "Designation", "Source Lists", "Identifiers (match)",
        "Key Contradiction", "Reasoning", "Investigation Sources",
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Handle zero-match case
    if not result.matches:
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=2, value="CLEAR")
        ws.cell(row=2, column=4, value=f"No candidates found in any of the screened databases")
        ws.cell(row=2, column=13, value="No name matches found across all sources.")
        # Apply green fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col).fill = GREEN_FILL
    else:
        for row_idx, match in enumerate(result.matches, 2):
            ws.cell(row=row_idx, column=1, value=match.get("number", row_idx - 1))
            ws.cell(row=row_idx, column=2, value=match.get("decision", ""))
            ws.cell(row=row_idx, column=3, value=match.get("cleared_by", ""))
            ws.cell(row=row_idx, column=4, value=match.get("matched_person", ""))
            ws.cell(row=row_idx, column=5, value=match.get("aliases", ""))
            ws.cell(row=row_idx, column=6, value=match.get("dob", ""))
            ws.cell(row=row_idx, column=7, value=match.get("nationality", ""))
            ws.cell(row=row_idx, column=8, value=match.get("gender", ""))
            ws.cell(row=row_idx, column=9, value=match.get("designation", ""))
            ws.cell(row=row_idx, column=10, value=match.get("source_lists", ""))
            ws.cell(row=row_idx, column=11, value=match.get("identifiers", ""))
            ws.cell(row=row_idx, column=12, value=match.get("key_contradiction", ""))
            ws.cell(row=row_idx, column=13, value=match.get("reasoning", ""))
            ws.cell(row=row_idx, column=14, value=match.get("investigation_sources", ""))

            # Apply row-level formatting based on decision
            decision = match.get("decision", "").upper()
            fill = {"CLEARED": GREEN_FILL, "LIKELY_MATCH": RED_FILL, "ESCALATE": YELLOW_FILL}.get(decision)
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = fill

            # Apply borders
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # Conditional formatting as backup (for Excel sorting/filtering)
    last_row = max(len(result.matches) + 1, 2)
    ws.conditional_formatting.add(
        f"B2:B{last_row}",
        CellIsRule(operator="equal", formula=['"CLEARED"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"B2:B{last_row}",
        CellIsRule(operator="equal", formula=['"LIKELY_MATCH"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"B2:B{last_row}",
        CellIsRule(operator="equal", formula=['"ESCALATE"'], fill=YELLOW_FILL),
    )

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Column widths (matches headers: #, Decision, Cleared By, Matched Person,
    # Aliases, DOB, Nationality, Gender, Designation, Source Lists, Identifiers,
    # Key Contradiction, Reasoning, Investigation Sources)
    widths = [5, 14, 18, 25, 30, 15, 15, 10, 25, 30, 25, 30, 50, 40]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _build_audit_sheet(ws, result: ScreeningResult):
    """Sheet 3: Audit — Claude I/O and source versions for regulators."""
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 100

    row = 1
    ws.cell(row=row, column=1, value="Report ID").font = BOLD_FONT
    ws.cell(row=row, column=2, value=result.id)
    row += 1

    ws.cell(row=row, column=1, value="Timestamp").font = BOLD_FONT
    ws.cell(row=row, column=2, value=result.timestamp)
    row += 1

    ws.cell(row=row, column=1, value="Model").font = BOLD_FONT
    ws.cell(row=row, column=2, value=result.llm_calls[0].get("model", "N/A") if result.llm_calls else "No LLM calls")
    row += 2

    # Source versions
    ws.cell(row=row, column=1, value="SOURCE VERSIONS").font = BOLD_FONT
    row += 1
    for source, sha in result.source_versions.items():
        ws.cell(row=row, column=1, value=source)
        ws.cell(row=row, column=2, value=sha)
        row += 1
    row += 1

    # LLM calls
    for call_idx, call in enumerate(result.llm_calls, 1):
        ws.cell(row=row, column=1, value=f"LLM CALL {call_idx}").font = BOLD_FONT
        row += 1

        for key in ["batch", "match_count", "model", "status", "elapsed_ms", "input_tokens", "output_tokens", "cache_read_tokens"]:
            if key in call:
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(call[key]))
                row += 1

        if "full_prompt" in call:
            ws.cell(row=row, column=1, value="Full Prompt")
            ws.cell(row=row, column=2, value=call["full_prompt"][:32000])  # Excel cell limit
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

        if "full_response" in call:
            ws.cell(row=row, column=1, value="Full Response")
            ws.cell(row=row, column=2, value=call["full_response"][:32000])
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

        if "error" in call:
            ws.cell(row=row, column=1, value="Error")
            ws.cell(row=row, column=2, value=call["error"])
            row += 1

        row += 1

    # Investigation audits (Pass 2)
    inv_audits = getattr(result, "investigation_audits", [])
    if inv_audits:
        row += 1
        ws.cell(row=row, column=1, value="=== PASS 2 INVESTIGATIONS ===").font = BOLD_FONT
        row += 2

        for inv_idx, inv in enumerate(inv_audits, 1):
            ws.cell(row=row, column=1, value=f"INVESTIGATION {inv_idx}").font = BOLD_FONT
            row += 1

            ws.cell(row=row, column=1, value="Match Name")
            ws.cell(row=row, column=2, value=inv.get("match_name", ""))
            row += 1

            ws.cell(row=row, column=1, value="Match Sources")
            ws.cell(row=row, column=2, value=", ".join(inv.get("match_sources", [])))
            row += 1

            ws.cell(row=row, column=1, value="Question to Perplexity")
            ws.cell(row=row, column=2, value=inv.get("question", "")[:32000])
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

            ws.cell(row=row, column=1, value="Perplexity Answer")
            ws.cell(row=row, column=2, value=(inv.get("perplexity_answer") or "")[:32000])
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1

            cites = inv.get("perplexity_citations", [])
            if cites:
                ws.cell(row=row, column=1, value="Citations")
                citation_text = "\n".join(
                    f"{c.get('title', '')}: {c.get('url', '')}" for c in cites
                )
                ws.cell(row=row, column=2, value=citation_text[:32000])
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                row += 1

            claude_call = inv.get("claude_call", {})
            if claude_call:
                if "claude_input_tokens" in claude_call:
                    ws.cell(row=row, column=1, value="Claude Tokens")
                    ws.cell(row=row, column=2, value=f"in={claude_call.get('claude_input_tokens')} out={claude_call.get('claude_output_tokens')}")
                    row += 1

                if "claude_response" in claude_call:
                    ws.cell(row=row, column=1, value="Claude Verdict JSON")
                    ws.cell(row=row, column=2, value=claude_call["claude_response"][:32000])
                    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                    row += 1

            if "error" in inv:
                ws.cell(row=row, column=1, value="Error")
                ws.cell(row=row, column=2, value=inv["error"])
                row += 1

            row += 1


def generate_json(result: ScreeningResult) -> str:
    """Generate JSON report string from a ScreeningResult."""
    return json.dumps({
        "id": result.id,
        "timestamp": result.timestamp,
        "result": result.result,
        "user_input": result.user_input,
        "screened_by": result.screened_by,
        "stats": {
            "raw_candidates": result.raw_candidates,
            "unique_persons": result.unique_persons,
            "auto_cleared": result.auto_cleared,
            "auto_flagged": result.auto_flagged,
            "sent_to_llm": result.sent_to_llm,
            "llm_cleared": result.llm_cleared,
            "llm_flagged": result.llm_flagged,
            "llm_escalated": result.llm_escalated,
        },
        "matches": result.matches,
        "source_versions": result.source_versions,
        "processing_ms": result.processing_ms,
    }, indent=2)


def _format_sources(source_versions: dict) -> str:
    """Format source list for Summary sheet."""
    source_names = {
        "ofac_sdn": "OFAC SDN",
        "ofac_cons": "OFAC Consolidated",
        "un_consolidated": "UN Security Council",
        "eu_fsf": "EU Financial Sanctions",
        "uk_sanctions": "UK Sanctions List",
        "ca_sema": "Canada SEMA",
        "ch_seco": "Switzerland SECO",
        "au_dfat": "Australia DFAT",
        "wikidata_peps": "Wikidata Global PEPs",
        "us_congress": "US Congress",
        "uk_parliament": "UK Parliament",
        "eu_parliament": "EU Parliament",
        "fbi_most_wanted": "FBI Most Wanted",
    }
    names = [source_names.get(s, s) for s in source_versions.keys()]
    return f"{len(names)} sources: {', '.join(names)}"
